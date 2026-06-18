from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.utils import timezone

from projects.models import Project, ProjectMember
from tasks.models import Task


# ══════════════════════════════════════════════════════════════
#  REPORTS & EXPORT  /reports/
# ══════════════════════════════════════════════════════════════

@login_required
def reports_view(request):
    today = timezone.localdate()

    # โปรเจคที่ user เข้าถึงได้
    member_pids = ProjectMember.objects.filter(user=request.user).values_list('project_id', flat=True)
    accessible  = Project.objects.filter(
        Q(owner=request.user) | Q(id__in=member_pids)
    ).distinct().order_by('name')

    # ── Filter params ──────────────────────────────────────────
    project_id = request.GET.get('project', '')
    status     = request.GET.get('status', '')
    priority   = request.GET.get('priority', '')
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')

    tasks = Task.objects.filter(
        Q(project__owner=request.user) | Q(project__id__in=member_pids)
    ).select_related('project', 'assigned_to', 'category').order_by('project__name', 'due_date')

    if project_id:
        tasks = tasks.filter(project_id=project_id)
    if status:
        tasks = tasks.filter(status=status)
    if priority:
        tasks = tasks.filter(priority=priority)
    if date_from:
        tasks = tasks.filter(due_date__gte=date_from)
    if date_to:
        tasks = tasks.filter(due_date__lte=date_to)

    total      = tasks.count()
    done_count = tasks.filter(status='done').count()
    over_count = tasks.exclude(status='done').filter(due_date__lt=today).count()
    comp_rate  = round(done_count / total * 100) if total else 0

    return render(request, 'projects/reports.html', {
        'tasks':        tasks[:200],   # preview สูงสุด 200
        'projects':     accessible,
        'total':        total,
        'done_count':   done_count,
        'over_count':   over_count,
        'comp_rate':    comp_rate,
        # filter state
        'sel_project':  project_id,
        'sel_status':   status,
        'sel_priority': priority,
        'sel_date_from': date_from,
        'sel_date_to':   date_to,
    })


# ══════════════════════════════════════════════════════════════
#  EXPORT EXCEL  /reports/export/excel/
# ══════════════════════════════════════════════════════════════

@login_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today      = timezone.localdate()
    member_pids = ProjectMember.objects.filter(user=request.user).values_list('project_id', flat=True)

    tasks = Task.objects.filter(
        Q(project__owner=request.user) | Q(project__id__in=member_pids)
    ).select_related('project', 'assigned_to', 'category')

    # Apply same filters as reports_view
    project_id = request.GET.get('project', '')
    status     = request.GET.get('status', '')
    priority   = request.GET.get('priority', '')
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')

    if project_id:
        tasks = tasks.filter(project_id=project_id)
    if status:
        tasks = tasks.filter(status=status)
    if priority:
        tasks = tasks.filter(priority=priority)
    if date_from:
        tasks = tasks.filter(due_date__gte=date_from)
    if date_to:
        tasks = tasks.filter(due_date__lte=date_to)

    tasks = tasks.order_by('project__name', 'due_date')

    wb = openpyxl.Workbook()

    # ── Sheet 1: Task Report ───────────────────────────────────
    ws = wb.active
    ws.title = 'Task Report'

    # Header style
    header_fill   = PatternFill('solid', fgColor='1a1a2e')
    header_font   = Font(bold=True, color='FFFFFF', size=11)
    center_align  = Alignment(horizontal='center', vertical='center')
    wrap_align    = Alignment(wrap_text=True, vertical='top')
    thin_border   = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    # Title row
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value       = f'KanFlow — Task Report  ({today.strftime("%d %b %Y")})'
    title_cell.font        = Font(bold=True, size=14, color='1a1a2e')
    title_cell.alignment   = center_align
    ws.row_dimensions[1].height = 30

    # Summary row
    ws.merge_cells('A2:I2')
    summary = ws['A2']
    summary.value     = (f'Total: {tasks.count()}  |  Done: {tasks.filter(status="done").count()}'
                         f'  |  Overdue: {tasks.exclude(status="done").filter(due_date__lt=today).count()}')
    summary.font      = Font(italic=True, color='666666', size=10)
    summary.alignment = center_align

    ws.row_dimensions[3].height = 22
    headers = ['#', 'Project', 'Task Title', 'Status', 'Priority', 'Assigned To', 'Category', 'Due Date', 'Overdue']
    col_widths = [5, 22, 35, 12, 12, 20, 16, 14, 10]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=3, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = w

    STATUS_COLORS = {'todo': 'E2E8F0', 'doing': 'FEF3C7', 'done': 'D1FAE5'}
    PRIORITY_COLORS = {'high': 'FEE2E2', 'medium': 'FEF3C7', 'low': 'F0FDF4'}

    for row_idx, task in enumerate(tasks, 4):
        is_overdue = task.due_date and task.due_date < today and task.status != 'done'
        row_fill   = PatternFill('solid', fgColor='FFF5F5') if is_overdue else None

        values = [
            row_idx - 3,
            task.project.name,
            task.title,
            task.get_status_display(),
            task.get_priority_display(),
            task.assigned_to.get_full_name() or task.assigned_to.username,
            task.category.name if task.category else '-',
            task.due_date.strftime('%d %b %Y') if task.due_date else '-',
            '⚠ Overdue' if is_overdue else '',
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border    = thin_border
            cell.alignment = wrap_align
            if row_fill:
                cell.fill = row_fill
            # Color status column
            if col == 4:
                color = STATUS_COLORS.get(task.status, 'FFFFFF')
                cell.fill = PatternFill('solid', fgColor=color)
            if col == 5:
                color = PRIORITY_COLORS.get(task.priority, 'FFFFFF')
                cell.fill = PatternFill('solid', fgColor=color)

        ws.row_dimensions[row_idx].height = 18

    # ── Sheet 2: Summary by Project ───────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 30
    for c in ['B','C','D','E','F']: ws2.column_dimensions[c].width = 14

    ws2.merge_cells('A1:F1')
    ws2['A1'].value     = 'Project Summary'
    ws2['A1'].font      = Font(bold=True, size=13, color='1a1a2e')
    ws2['A1'].alignment = center_align

    h2 = ['Project', 'Total', 'Done', 'Doing', 'Todo', 'Completion %']
    for col, h in enumerate(h2, 1):
        cell           = ws2.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    from projects.models import Project as ProjectModel
    member_pids2 = ProjectMember.objects.filter(user=request.user).values_list('project_id', flat=True)
    proj_qs = ProjectModel.objects.filter(
        Q(owner=request.user) | Q(id__in=member_pids2)
    ).distinct()
    if project_id:
        proj_qs = proj_qs.filter(id=project_id)

    for ri, p in enumerate(proj_qs, 3):
        pt = tasks.filter(project=p)
        total_p = pt.count()
        done_p  = pt.filter(status='done').count()
        comp_p  = round(done_p / total_p * 100) if total_p else 0
        row_vals = [p.name, total_p, done_p, pt.filter(status='doing').count(),
                    pt.filter(status='todo').count(), f'{comp_p}%']
        for col, v in enumerate(row_vals, 1):
            cell           = ws2.cell(row=ri, column=col, value=v)
            cell.border    = thin_border
            cell.alignment = center_align if col > 1 else wrap_align

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="kanflow_report_{today}.xlsx"'
    wb.save(response)
    return response


# ══════════════════════════════════════════════════════════════
#  EXPORT PDF  /reports/export/pdf/
# ══════════════════════════════════════════════════════════════

@login_required
def export_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io, os

    today      = timezone.localdate()
    member_pids = ProjectMember.objects.filter(user=request.user).values_list('project_id', flat=True)

    tasks = Task.objects.filter(
        Q(project__owner=request.user) | Q(project__id__in=member_pids)
    ).select_related('project', 'assigned_to', 'category')

    project_id = request.GET.get('project', '')
    status     = request.GET.get('status', '')
    priority   = request.GET.get('priority', '')
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')

    if project_id:
        tasks = tasks.filter(project_id=project_id)
    if status:
        tasks = tasks.filter(status=status)
    if priority:
        tasks = tasks.filter(priority=priority)
    if date_from:
        tasks = tasks.filter(due_date__gte=date_from)
    if date_to:
        tasks = tasks.filter(due_date__lte=date_to)

    tasks = tasks.order_by('project__name', 'due_date')[:500]

    # ── Register Thai font (ถ้ามี) ────────────────────────────
    # Use THSarabun font for Thai character support
    FONT_NAME = 'THSarabun'
    
    try:
        import os
        from django.conf import settings
        
        # Register THSarabun font with absolute path
        font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'THSarabunNew.ttf')
        
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont('THSarabun', font_path))
            print(f"Successfully registered THSarabun font from: {font_path}")
        else:
            print(f"THSarabun font not found at: {font_path}")
            print("Using fallback font: Helvetica")
            FONT_NAME = 'Helvetica'
            
    except Exception as e:
        print(f"Font registration error: {e}")
        FONT_NAME = 'Helvetica'

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    COLOR_DARK   = colors.HexColor('#1a1a2e')
    COLOR_ACCENT = colors.HexColor('#6366f1')
    COLOR_LIGHT  = colors.HexColor('#f8fafc')
    COLOR_BORDER = colors.HexColor('#e2e8f0')

    styles    = getSampleStyleSheet()
    style_h1  = ParagraphStyle('H1', fontName=FONT_NAME, fontSize=20, textColor=COLOR_DARK,
                                spaceAfter=4, leading=24, alignment=TA_LEFT)
    style_sub = ParagraphStyle('Sub', fontName=FONT_NAME, fontSize=10, textColor=colors.HexColor('#64748b'),
                                spaceAfter=12)
    style_cell = ParagraphStyle('Cell', fontName=FONT_NAME, fontSize=8, leading=12)
    style_hdr  = ParagraphStyle('Hdr', fontName=FONT_NAME, fontSize=9, textColor=colors.white,
                                 leading=12, alignment=TA_CENTER)

    total_count = len(tasks)
    done_count  = sum(1 for t in tasks if t.status == 'done')
    over_count  = sum(1 for t in tasks if t.due_date and t.due_date < today and t.status != 'done')

    elements = []

    # Title
    elements.append(Paragraph('KanFlow — Task Report', style_h1))
    elements.append(Paragraph(
        f'Generated: {today.strftime("%d %B %Y")}  |  '
        f'Total: {total_count}  |  Done: {done_count}  |  Overdue: {over_count}',
        style_sub
    ))
    elements.append(HRFlowable(width='100%', thickness=2, color=COLOR_ACCENT, spaceAfter=10))

    # Table
    col_widths_pdf = [1*cm, 5*cm, 7.5*cm, 2.8*cm, 2.8*cm, 4*cm, 3.5*cm, 3*cm]
    header_row = [
        Paragraph('#',          style_hdr),
        Paragraph('Project',    style_hdr),
        Paragraph('Task Title', style_hdr),
        Paragraph('Status',     style_hdr),
        Paragraph('Priority',   style_hdr),
        Paragraph('Assigned To',style_hdr),
        Paragraph('Due Date',   style_hdr),
        Paragraph('Category',   style_hdr),
    ]

    STATUS_COLORS_PDF = {
        'todo':  colors.HexColor('#CBD5E1'),
        'doing': colors.HexColor('#FDE68A'),
        'done':  colors.HexColor('#6EE7B7'),
    }
    PRIORITY_COLORS_PDF = {
        'high':   colors.HexColor('#FCA5A5'),
        'medium': colors.HexColor('#FDE68A'),
        'low':    colors.HexColor('#A7F3D0'),
    }

    data = [header_row]
    row_styles = []

    for idx, task in enumerate(tasks, 1):
        is_overdue = task.due_date and task.due_date < today and task.status != 'done'
        row_bg = colors.HexColor('#FFF5F5') if is_overdue else (
            colors.white if idx % 2 == 0 else COLOR_LIGHT
        )
        row_styles.append(('BACKGROUND', (0, idx), (-1, idx), row_bg))

        # Status badge color
        row_styles.append(('BACKGROUND', (3, idx), (3, idx),
                            STATUS_COLORS_PDF.get(task.status, colors.white)))
        row_styles.append(('BACKGROUND', (4, idx), (4, idx),
                            PRIORITY_COLORS_PDF.get(task.priority, colors.white)))

        data.append([
            Paragraph(str(idx),                                              style_cell),
            Paragraph(task.project.name,                                    style_cell),
            Paragraph(task.title,                                         style_cell),
            Paragraph(task.get_status_display(),                             style_cell),
            Paragraph(task.get_priority_display(),                           style_cell),
            Paragraph(task.assigned_to.get_full_name() or task.assigned_to.username, style_cell),
            Paragraph(task.due_date.strftime('%d %b %Y') if task.due_date else '-', style_cell),
            Paragraph(task.category.name[:20] if task.category else '-',    style_cell),
        ])

    base_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), COLOR_DARK),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_LIGHT]),
        ('GRID',       (0, 0), (-1, -1), 0.5, COLOR_BORDER),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
    ] + row_styles)

    tbl = Table(data, colWidths=col_widths_pdf, repeatRows=1)
    tbl.setStyle(base_style)
    elements.append(tbl)

    # Footer note
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(
        f'KanFlow — Exported by {request.user.username}  |  {today}',
        ParagraphStyle('Footer', fontName=FONT_NAME, fontSize=8,
                       textColor=colors.HexColor('#94a3b8'), alignment=TA_RIGHT)
    ))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="kanflow_report_{today}.pdf"'
    return response
